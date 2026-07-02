import io

from django.http import HttpResponse
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import F, Sum
from django.utils import timezone

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm

from catalog.models import Product
from inventory.models import StockMovement
from orders.models import Customer, Order, OrderItem
from procurement.models import Purchase

from .services.eoq import calculate_eoq, get_annual_demand
from .services.abc import build_abc_report
from .services.trends import build_sales_trend

HEADER_FILL = PatternFill("solid", fgColor="7B3F1E")
HEADER_FONT = Font(bold=True, color="FFFFFF")


@login_required
def dashboard(request):
    from datetime import timedelta
    from decimal import Decimal

    today = timezone.localdate()
    this_month_start = today.replace(day=1)
    last_month_end = this_month_start - timedelta(days=1)
    last_month_start = last_month_end.replace(day=1)

    # ── Core counts ──────────────────────────────────────────────────────────
    total_products  = Product.objects.count()
    total_orders    = Order.objects.filter(status="CONFIRMED").count()
    total_purchases = Purchase.objects.filter(status="RECEIVED").count()
    total_stock     = Product.objects.aggregate(total=Sum("current_stock"))["total"] or 0
    low_stock_products = Product.objects.filter(current_stock__lt=F("reorder_point"))
    draft_orders_count = Order.objects.filter(status="DRAFT").count()

    # ── Today ─────────────────────────────────────────────────────────────────
    today_confirmed = (
        Order.objects
        .filter(status="CONFIRMED", created_at__date=today)
        .prefetch_related("items__product")
    )
    today_order_count = today_confirmed.count()
    today_revenue = sum(
        item.quantity * item.product.selling_price
        for order in today_confirmed
        for item in order.items.all()
    ) or Decimal("0")

    # ── This month vs last month ──────────────────────────────────────────────
    this_month_items = OrderItem.objects.filter(
        order__status="CONFIRMED",
        order__created_at__date__gte=this_month_start,
    ).select_related("product")
    this_month_revenue = sum(
        i.quantity * i.product.selling_price for i in this_month_items
    ) or Decimal("0")

    last_month_items = OrderItem.objects.filter(
        order__status="CONFIRMED",
        order__created_at__date__gte=last_month_start,
        order__created_at__date__lte=last_month_end,
    ).select_related("product")
    last_month_revenue = sum(
        i.quantity * i.product.selling_price for i in last_month_items
    ) or Decimal("0")

    if last_month_revenue > 0:
        month_change = round(
            float((this_month_revenue - last_month_revenue) / last_month_revenue * 100), 1
        )
    else:
        month_change = None

    # ── Recent orders (last 5, any status) ───────────────────────────────────
    recent_orders = []
    for order in (
        Order.objects
        .select_related("customer", "created_by")
        .prefetch_related("items__product")
        .order_by("-id")[:5]
    ):
        total = sum(i.quantity * i.product.selling_price for i in order.items.all())
        recent_orders.append({"order": order, "total": total})

    # ── Charts ────────────────────────────────────────────────────────────────
    top_products = (
        OrderItem.objects
        .filter(order__status="CONFIRMED")
        .values("product__name")
        .annotate(qty=Sum("quantity"))
        .order_by("-qty")[:5]
    )
    top_labels = [x["product__name"] for x in top_products]
    top_qty    = [x["qty"] for x in top_products]

    trend_rows   = build_sales_trend()
    trend_labels = [r["month"] for r in trend_rows]
    trend_values = [float(r["revenue"]) for r in trend_rows]

    return render(request, "analytics/dashboard.html", {
        "total_products":      total_products,
        "total_orders":        total_orders,
        "total_purchases":     total_purchases,
        "total_stock":         total_stock,
        "low_stock_products":  low_stock_products,
        "draft_orders_count":  draft_orders_count,
        "today_order_count":   today_order_count,
        "today_revenue":       today_revenue,
        "this_month_revenue":  this_month_revenue,
        "last_month_revenue":  last_month_revenue,
        "month_change":        month_change,
        "recent_orders":       recent_orders,
        "top_labels":          top_labels,
        "top_qty":             top_qty,
        "trend_labels":        trend_labels,
        "trend_values":        trend_values,
    })


@login_required
def eoq_report(request):
    products = Product.objects.all()

    report_data = []
    for product in products:
        demand, demand_source = get_annual_demand(product)
        eoq = calculate_eoq(product, annual_demand=demand)
        report_data.append({
            "product": product,
            "eoq": eoq,
            "annual_demand": demand,
            "demand_source": demand_source,
        })

    return render(request, "analytics/eoq_report.html", {
        "report_data": report_data
    })


@login_required
def abc_report(request):
    rows, total_value = build_abc_report()

    labels = [r["product"].name for r in rows]
    values = [float(r["sales_value"]) for r in rows]

    return render(request, "analytics/abc_report.html", {
        "rows": rows,
        "total_value": total_value,
        "labels": labels,
        "values": values,
    })


@login_required
def sales_trend_report(request):
    rows = build_sales_trend()

    labels = [r["month"] for r in rows]
    revenue = [float(r["revenue"]) for r in rows]

    return render(request, "analytics/sales_trend.html", {
        "rows": rows,
        "labels": labels,
        "revenue": revenue,
    })


# ── Helpers ──────────────────────────────────────────────────────────────────

def _style_header_row(ws, row_num, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _pdf_table_style():
    return TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#7B3F1E")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9F3EE")]),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#DDCCBB")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ])


# ── Inventory export ──────────────────────────────────────────────────────────

@login_required
def export_inventory_excel(request):
    products = Product.objects.select_related("category").order_by("name")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory"
    headers = ["#", "Product", "Category", "Selling Price (NPR)", "Unit Cost (NPR)", "Current Stock", "Status"]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))
    for i, p in enumerate(products, 1):
        if p.current_stock < p.reorder_point:
            status = "Low Stock"
        elif p.current_stock < p.reorder_point + 15:
            status = "Medium"
        else:
            status = "OK"
        ws.append([i, p.name, p.category.name, float(p.selling_price), float(p.unit_cost), p.current_stock, status])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(len(str(col[0].value or "")), 12) + 4
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(buf, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="inventory_{timezone.now():%Y%m%d}.xlsx"'
    return response


@login_required
def export_inventory_pdf(request):
    products = Product.objects.select_related("category").order_by("name")
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=1.5*cm, rightMargin=1.5*cm, topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph("Lophoro IMS — Inventory Report", styles["Title"]),
        Paragraph(f"Generated: {timezone.now():%d %b %Y}", styles["Normal"]),
        Spacer(1, 0.4*cm),
    ]
    data = [["#", "Product", "Category", "Selling Price", "Unit Cost", "Stock", "Status"]]
    for i, p in enumerate(products, 1):
        status = "Low" if p.current_stock < p.reorder_point else ("Medium" if p.current_stock < p.reorder_point + 15 else "OK")
        data.append([i, p.name, p.category.name, f"NPR {p.selling_price}", f"NPR {p.unit_cost}", p.current_stock, status])
    t = Table(data, repeatRows=1)
    t.setStyle(_pdf_table_style())
    elements.append(t)
    doc.build(elements)
    buf.seek(0)
    response = HttpResponse(buf, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="inventory_{timezone.now():%Y%m%d}.pdf"'
    return response


# ── Orders export ─────────────────────────────────────────────────────────────

@login_required
def export_orders_excel(request):
    orders = Order.objects.select_related("created_by", "customer").prefetch_related("items__product").order_by("-id")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orders"
    headers = ["Order #", "Customer", "Created By", "Date", "Status", "Items", "Total (NPR)"]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))
    for o in orders:
        items_str = ", ".join(f"{i.product.name} x{i.quantity}" for i in o.items.all())
        total = sum(i.quantity * i.product.selling_price for i in o.items.all())
        ws.append([
            o.id,
            o.customer.name if o.customer else "—",
            o.created_by.username,
            o.created_at.strftime("%Y-%m-%d %H:%M"),
            o.status,
            items_str,
            float(total),
        ])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(len(str(col[0].value or "")), 12) + 4
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(buf, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="orders_{timezone.now():%Y%m%d}.xlsx"'
    return response


@login_required
def export_orders_pdf(request):
    orders = Order.objects.select_related("created_by", "customer").prefetch_related("items__product").order_by("-id")
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=1.5*cm, rightMargin=1.5*cm, topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph("Lophoro IMS — Orders Report", styles["Title"]),
        Paragraph(f"Generated: {timezone.now():%d %b %Y}", styles["Normal"]),
        Spacer(1, 0.4*cm),
    ]
    data = [["Order #", "Customer", "Created By", "Date", "Status", "Total (NPR)"]]
    for o in orders:
        total = sum(i.quantity * i.product.selling_price for i in o.items.all())
        data.append([
            f"#{o.id}",
            o.customer.name if o.customer else "—",
            o.created_by.username,
            o.created_at.strftime("%Y-%m-%d"),
            o.status,
            f"NPR {total:,.2f}",
        ])
    t = Table(data, repeatRows=1)
    t.setStyle(_pdf_table_style())
    elements.append(t)
    doc.build(elements)
    buf.seek(0)
    response = HttpResponse(buf, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="orders_{timezone.now():%Y%m%d}.pdf"'
    return response


# ── Purchases export ──────────────────────────────────────────────────────────

@login_required
def export_purchases_excel(request):
    purchases = Purchase.objects.select_related("supplier", "created_by").prefetch_related("items__product").order_by("-id")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Purchases"
    headers = ["Purchase #", "Supplier", "Created By", "Date", "Status", "Items", "Total Cost (NPR)"]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))
    for p in purchases:
        items_str = ", ".join(f"{i.product.name} x{i.quantity}" for i in p.items.all())
        total = sum(i.quantity * i.unit_cost for i in p.items.all())
        ws.append([
            p.id,
            p.supplier.name,
            p.created_by.username,
            p.created_at.strftime("%Y-%m-%d %H:%M"),
            p.status,
            items_str,
            float(total),
        ])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(len(str(col[0].value or "")), 12) + 4
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(buf, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="purchases_{timezone.now():%Y%m%d}.xlsx"'
    return response


@login_required
def export_purchases_pdf(request):
    purchases = Purchase.objects.select_related("supplier", "created_by").prefetch_related("items__product").order_by("-id")
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=1.5*cm, rightMargin=1.5*cm, topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph("Lophoro IMS — Purchases Report", styles["Title"]),
        Paragraph(f"Generated: {timezone.now():%d %b %Y}", styles["Normal"]),
        Spacer(1, 0.4*cm),
    ]
    data = [["Purchase #", "Supplier", "Created By", "Date", "Status", "Total Cost (NPR)"]]
    for p in purchases:
        total = sum(i.quantity * i.unit_cost for i in p.items.all())
        data.append([
            f"#{p.id}",
            p.supplier.name,
            p.created_by.username,
            p.created_at.strftime("%Y-%m-%d"),
            p.status,
            f"NPR {total:,.2f}",
        ])
    t = Table(data, repeatRows=1)
    t.setStyle(_pdf_table_style())
    elements.append(t)
    doc.build(elements)
    buf.seek(0)
    response = HttpResponse(buf, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="purchases_{timezone.now():%Y%m%d}.pdf"'
    return response


# ── Customers export ──────────────────────────────────────────────────────────

@login_required
def export_customers_excel(request):
    customers = Customer.objects.prefetch_related("orders").order_by("name")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customers"
    headers = ["#", "Name", "Phone", "Email", "Address", "Total Orders", "Joined"]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))
    for c in customers:
        ws.append([
            c.id,
            c.name,
            c.phone or "—",
            c.email or "—",
            c.address or "—",
            c.orders.count(),
            c.created_at.strftime("%Y-%m-%d"),
        ])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(len(str(col[0].value or "")), 12) + 4
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(buf, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="customers_{timezone.now():%Y%m%d}.xlsx"'
    return response


@login_required
def export_customers_pdf(request):
    customers = Customer.objects.prefetch_related("orders").order_by("name")
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=1.5*cm, rightMargin=1.5*cm, topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph("Lophoro IMS — Customer List", styles["Title"]),
        Paragraph(f"Generated: {timezone.now():%d %b %Y}", styles["Normal"]),
        Spacer(1, 0.4*cm),
    ]
    data = [["#", "Name", "Phone", "Email", "Address", "Orders"]]
    for c in customers:
        data.append([
            c.id,
            c.name,
            c.phone or "—",
            c.email or "—",
            c.address or "—",
            c.orders.count(),
        ])
    t = Table(data, repeatRows=1, colWidths=[1*cm, 4*cm, 3*cm, 5*cm, 5*cm, 1.5*cm])
    t.setStyle(_pdf_table_style())
    elements.append(t)
    doc.build(elements)
    buf.seek(0)
    response = HttpResponse(buf, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="customers_{timezone.now():%Y%m%d}.pdf"'
    return response